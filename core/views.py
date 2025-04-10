import csv
from openpyxl import load_workbook
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Sum, Count, Case, When, IntegerField
from django.core.paginator import Paginator
from django.contrib.auth import authenticate, login, logout
from django.forms import inlineformset_factory

from .models import Medecin, Commentaire, Aspect, Aborder
from .forms import MedecinForm, CommentaireForm, CustomLoginForm, CommentFileUploadForm, DoctorFileUploadForm, AborderForm, AspectForm
from django.forms import inlineformset_factory
from .forms import AborderForm  
import json
from django.http import JsonResponse
from .models import Medecin, Quartier
from django.core.exceptions import ValidationError


# ------------------ CRUD Médecin ------------------

def doctor_list(request):
    search_query = request.GET.get('search', '')
    doctors_qs = Medecin.objects.all().order_by('nom')
    if search_query:
        doctors_qs = doctors_qs.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(specialite__icontains=search_query) |
            Q(quartier__nom__icontains=search_query) |
            Q(quartier__ville__icontains=search_query)
        )
    paginator = Paginator(doctors_qs, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'core/doctor_list.html', context)


def doctor_create(request):
    if request.method == 'POST':
        form = MedecinForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('doctor_list')
    else:
        form = MedecinForm()
    return render(request, 'core/doctor_form.html', {'form': form, 'title': 'Add Doctor'})

def doctor_edit(request, pk):
    doctor = get_object_or_404(Medecin, pk=pk)
    if request.method == 'POST':
        form = MedecinForm(request.POST, instance=doctor)
        if form.is_valid():
            form.save()
            return redirect('doctor_list')
    else:
        form = MedecinForm(instance=doctor)
    return render(request, 'core/doctor_form.html', {'form': form, 'title': 'Edit Doctor'})

def doctor_delete(request, pk):
    doctor = get_object_or_404(Medecin, pk=pk)
    doctor_name = f"{doctor.prenom} {doctor.nom}"
    if request.method == 'POST':
        doctor.delete()
        messages.success(request, f"Doctor {doctor_name} has been deleted successfully.")
        return redirect('doctor_list')
    return redirect('doctor_list')

# ------------------ Gestion des Commentaires ------------------

def doctor_comments(request, pk):
    doctor = get_object_or_404(Medecin, pk=pk)
    comments = Commentaire.objects.filter(medecin=doctor).order_by('-date')

    aborders = Aborder.objects.filter(commentaire__medecin=doctor).annotate(
        numeric_polarity=Case(
            When(polarite='Positif', then=1),
            When(polarite='Negatif', then=-1),
            default=0,
            output_field=IntegerField()
        )
    )
    grouped_stats = aborders.values('aspect__nom_aspect').annotate(
        sum_score=Sum('numeric_polarity'),
        count=Count('id'),
    )
    stats_list = []
    for row in grouped_stats:
        aspect_name = row['aspect__nom_aspect']
        sum_score = row['sum_score'] or 0
        count = row['count'] or 0
        avg_score = round(sum_score / count, 2) if count > 0 else 0
        stats_list.append({
            'aspect': aspect_name,
            'sum_score': sum_score,
            'count': count,
            'avg_score': avg_score,
        })

    context = {
        'doctor': doctor,
        'comments': comments,
        'stats_list': stats_list,
    }
    return render(request, 'core/doctor_comments.html', context)

def comment_create_for_doctor(request, pk):
    doctor = get_object_or_404(Medecin, pk=pk)
    if request.method == 'POST':
        form = CommentaireForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.medecin = doctor
            comment.save()
            # Laisser l'administrateur saisir manuellement les aspects ensuite
            return redirect('doctor_comments', pk=doctor.pk)
    else:
        form = CommentaireForm()
    return render(request, 'core/comment_form.html', {'form': form, 'doctor': doctor})

def comment_edit(request, pk):
    comment = get_object_or_404(Commentaire, pk=pk)
    if request.method == 'POST':
        form = CommentaireForm(request.POST, instance=comment)
        if form.is_valid():
            form.save()
            return redirect('doctor_comments', pk=comment.medecin.pk)
    else:
        form = CommentaireForm(instance=comment)
    return render(request, 'core/comment_form.html', {'form': form, 'doctor': comment.medecin, 'title': 'Edit Comment'})

def comment_delete(request, pk):
    comment = get_object_or_404(Commentaire, pk=pk)
    if request.method == 'POST':
        doc_pk = comment.medecin.pk
        comment.delete()
        return redirect('doctor_comments', pk=doc_pk)
    return render(request, 'core/comment_confirm_delete.html', {'comment': comment})

# ------------------ CRUD Aspects Manuels (Aborder) ------------------

def annotate_comment(request, comment_id):
    """
    Affiche la page pour que l'administrateur saisisse manuellement
    les aspects et leur polarité pour un commentaire via un formset.
    """
    comment = get_object_or_404(Commentaire, pk=comment_id)
    AborderFormSet = inlineformset_factory(
        Commentaire, Aborder,
        form=AborderForm,
        fields=('aspect', 'polarite'),
        extra=1,
        can_delete=True
    )
    if request.method == 'POST':
        formset = AborderFormSet(request.POST, instance=comment)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Aspects updated successfully.")
            return redirect('comment_analyse', comment_id=comment.pk)
    else:
        formset = AborderFormSet(instance=comment)
    context = {
        'comment': comment,
        'formset': formset,
    }
    return render(request, 'core/annotate_comment.html', context)

def comment_analyse(request, comment_id):
    """
    Affiche l'analyse d'un commentaire : liste des aspects associés et leurs polarités.
    """
    comment = get_object_or_404(Commentaire, pk=comment_id)
    aspects_associes = Aborder.objects.select_related('aspect').filter(commentaire=comment)
    context = {
        'comment': comment,
        'aspects_associes': aspects_associes,
    }
    return render(request, 'core/comment_analyse.html', context)

def aborder_add(request, comment_id):
    """
    Permet d'ajouter un nouvel aspect (Aborder) à un commentaire via un formulaire.
    Empêche les doublons d'aspect pour un même commentaire.
    """
    comment = get_object_or_404(Commentaire, pk=comment_id)
    
    if request.method == 'POST':
        form = AborderForm(request.POST, commentaire=comment)
        if form.is_valid():
            aborder = form.save(commit=False)
            aborder.commentaire = comment
            aborder.save()
            messages.success(request, "Aspect ajouté avec succès.")
            return redirect('comment_analyse', comment_id=comment.pk)
    else:
        form = AborderForm(commentaire=comment)

    return render(request, 'core/aborder_add.html', {'form': form, 'comment': comment})


def aborder_edit(request, aborder_id):
    """
    Permet d'éditer un aspect existant.
    """
    aborder = get_object_or_404(Aborder, pk=aborder_id)
    comment = aborder.commentaire
    if request.method == 'POST':
        form = AborderForm(request.POST, instance=aborder)
        if form.is_valid():
            form.save()
            messages.success(request, "Aspect updated successfully.")
            return redirect('comment_analyse', comment_id=comment.pk)
    else:
        form = AborderForm(instance=aborder)
    return render(request, 'core/aborder_edit.html', {'form': form, 'aborder': aborder, 'comment': comment})

def aborder_delete(request, aborder_id):
    """
    Permet de supprimer un aspect (Aborder) après confirmation.
    """
    aborder = get_object_or_404(Aborder, pk=aborder_id)
    comment = aborder.commentaire
    if request.method == 'POST':
        aborder.delete()
        messages.success(request, "Aspect deleted successfully.")
        return redirect('comment_analyse', comment_id=comment.pk)
    return render(request, 'core/aborder_delete.html', {'aborder': aborder, 'comment': comment})

# ------------------ Import de Commentaires ------------------

def comment_import_for_doctor(request, pk):
    doctor = get_object_or_404(Medecin, pk=pk)
    if request.method == 'POST':
        form = CommentFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            filename = uploaded_file.name.lower()
            if filename.endswith('.csv'):
                handle_csv_comments(uploaded_file, doctor)
            elif filename.endswith('.xlsx'):
                handle_excel_comments(uploaded_file, doctor)
            else:
                form.add_error('file', "Le fichier doit être un CSV ou un fichier Excel (.xlsx).")
                return render(request, 'core/comment_import.html', {'form': form, 'doctor': doctor})
            messages.success(request, "Comments imported successfully.")
            return redirect('doctor_comments', pk=doctor.pk)
    else:
        form = CommentFileUploadForm()
    return render(request, 'core/comment_import.html', {'form': form, 'doctor': doctor})

def handle_csv_comments(uploaded_file, doctor):
    data = uploaded_file.read().decode('utf-8').splitlines()
    reader = csv.DictReader(data)
    for row in reader:
        contenu = row.get('contenu', '')
        if contenu.strip():
            comment = Commentaire.objects.create(medecin=doctor, contenu=contenu)
            # Ici, vous ne déclenchez pas l'analyse automatique
            # L'administrateur saisira manuellement les aspects

def handle_excel_comments(uploaded_file, doctor):
    wb = load_workbook(uploaded_file)
    sheet = wb.active
    headers = [cell.value.lower().strip() if cell.value else '' for cell in sheet[1]]
    if 'contenu' not in headers:
        return
    contenu_index = headers.index('contenu') + 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        contenu = row[contenu_index - 1]
        if contenu and isinstance(contenu, str) and contenu.strip():
            comment = Commentaire.objects.create(medecin=doctor, contenu=contenu)
            # Pas d'analyse automatique

# ------------------ Import de Médecins ------------------

from django.contrib import messages
from django.shortcuts import render, redirect
import csv
from openpyxl import load_workbook
from .forms import DoctorFileUploadForm
from .models import Medecin, Quartier

def import_doctors(request):
    if request.method == 'POST':
        form = DoctorFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            filename = uploaded_file.name.lower()

            REQUIRED_COLUMNS = ['nom', 'prenom', 'specialite', 'email', 'telephone', 'nom_quartier', 'ville_quartier']

            # Fonction utilitaire pour convertir proprement en string
            def safe_str(value):
                return str(value).strip() if value is not None else ''

            # Fonction pour récupérer ou créer un quartier
            def get_or_create_quartier(nom_quartier, ville_quartier):
                if nom_quartier and ville_quartier:
                    return Quartier.objects.get_or_create(
                        nom=safe_str(nom_quartier),
                        ville=safe_str(ville_quartier)
                    )[0]
                return None

            doctors_added = 0
            doctors_skipped = 0

            # Traitement CSV
            if filename.endswith('.csv'):
                data = uploaded_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(data)
                csv_columns = [col.lower().strip() for col in reader.fieldnames]

                missing_columns = [col for col in REQUIRED_COLUMNS if col not in csv_columns]
                if missing_columns:
                    form.add_error('file', f"Colonnes manquantes dans le fichier CSV : {', '.join(missing_columns)}")
                    return render(request, 'core/doctor_import.html', {'form': form})

                for row in reader:
                    quartier = get_or_create_quartier(row['nom_quartier'], row['ville_quartier'])

                    nom = safe_str(row['nom'])
                    prenom = safe_str(row['prenom'])
                    email = safe_str(row['email'])

                    medecin_exists = Medecin.objects.filter(
                        nom=nom,
                        prenom=prenom,
                        email=email
                    ).exists()

                    if not medecin_exists:
                        Medecin.objects.create(
                            nom=nom,
                            prenom=prenom,
                            specialite=safe_str(row['specialite']),
                            email=email,
                            telephone=safe_str(row['telephone']),
                            quartier=quartier
                        )
                        doctors_added += 1
                    else:
                        doctors_skipped += 1

            # Traitement Excel
            elif filename.endswith('.xlsx'):
                wb = load_workbook(uploaded_file)
                sheet = wb.active
                headers = [safe_str(cell.value).lower() for cell in sheet[1]]

                missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
                if missing_columns:
                    form.add_error('file', f"Colonnes manquantes dans le fichier Excel : {', '.join(missing_columns)}")
                    return render(request, 'core/doctor_import.html', {'form': form})

                # Obtenir les index des colonnes
                index_map = {col: headers.index(col) for col in REQUIRED_COLUMNS}

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    nom = safe_str(row[index_map['nom']])
                    prenom = safe_str(row[index_map['prenom']])
                    email = safe_str(row[index_map['email']])
                    specialite = safe_str(row[index_map['specialite']])
                    telephone = safe_str(row[index_map['telephone']])
                    quartier_nom = safe_str(row[index_map['nom_quartier']])
                    quartier_ville = safe_str(row[index_map['ville_quartier']])
                    quartier = get_or_create_quartier(quartier_nom, quartier_ville)

                    medecin_exists = Medecin.objects.filter(
                        nom=nom,
                        prenom=prenom,
                        email=email
                    ).exists()

                    if not medecin_exists:
                        Medecin.objects.create(
                            nom=nom,
                            prenom=prenom,
                            specialite=specialite,
                            email=email,
                            telephone=telephone,
                            quartier=quartier
                        )
                        doctors_added += 1
                    else:
                        doctors_skipped += 1

            else:
                form.add_error('file', "Le fichier doit être au format CSV ou Excel (.xlsx).")
                return render(request, 'core/doctor_import.html', {'form': form})

            # Messages finaux
            messages.success(request, f"{doctors_added} médecins ont été ajoutés avec succès.")
            if doctors_skipped > 0:
                messages.info(request, f"{doctors_skipped} médecins ont été ignorés car ils existaient déjà.")

            return redirect('doctor_list')

    else:
        form = DoctorFileUploadForm()

    return render(request, 'core/doctor_import.html', {'form': form})


# ------------------ Sélection d'un médecin pour voir ses commentaires ------------------

def select_doctor_for_comments(request):
    search_query = request.GET.get('search', '')
    if search_query:
        doctors = Medecin.objects.filter(nom__icontains=search_query)
    else:
        doctors = Medecin.objects.all()
    context = {
        'doctors': doctors,
        'search_query': search_query,
    }
    return render(request, 'core/comment_list_doctors.html', context)

# ------------------ Authentification ------------------

def login_custom(request):
    error_message = None
    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return redirect('dashboard')
            else:
                error_message = "Nom d'utilisateur ou mot de passe invalide."
    else:
        form = CustomLoginForm()
    return render(request, 'core/login.html', {'form': form, 'error_message': error_message})

def logout_view(request):
    logout(request)
    return redirect('login')

def dashboard(request):
    total_doctors = Medecin.objects.count()
    total_comments = Commentaire.objects.count()
    context = {
        'total_doctors': total_doctors,
        'total_comments': total_comments,
    }
    return render(request, 'core/dashboard.html', context)

def annotate_comment(request, comment_id):
    comment = get_object_or_404(Commentaire, pk=comment_id)

    AborderFormSet = inlineformset_factory(
        Commentaire,
        Aborder,
        form=AborderForm,             # <-- IMPORTANT
        fields=('aspect', 'polarite'),
        extra=3,                      # 3 lignes vierges par défaut
        can_delete=True
    )

    if request.method == 'POST':
        formset = AborderFormSet(request.POST, instance=comment)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Aspects mis à jour avec succès.")
            return redirect('comment_analyse', comment_id=comment.pk)
    else:
        formset = AborderFormSet(instance=comment)

    # Pour afficher la liste déjà associée
    aspects_associes = Aborder.objects.filter(commentaire=comment)

    context = {
        'comment': comment,
        'formset': formset,
        'aspects_associes': aspects_associes,
    }
    return render(request, 'core/annotate_comment.html', context)

def doctor_list(request):
    doctors = Medecin.objects.all()
    return render(request, 'core/doctor_list.html', {'doctors': doctors})

def aspect_create_ajax(request):
    """
    Crée un nouvel aspect à partir d'une requête AJAX.
    Renvoie en JSON l'ID et le nom de l'aspect créé.
    """
    if request.method == 'POST' and request.is_ajax():
        form = AspectForm(request.POST)
        if form.is_valid():
            aspect = form.save()
            return JsonResponse({'success': True, 'id': aspect.id, 'text': aspect.nom_aspect})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    return JsonResponse({'success': False, 'error': 'Requête invalide'})